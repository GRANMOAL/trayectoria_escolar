import json as _json
import io
import base64
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from matplotlib.colors import LinearSegmentedColormap
from openpyxl.utils import get_column_letter

from .models import Semestre, Grupo, Alumno, Asignatura, Calificacion

# ── Constante global ──────────────────────────────────────────────────────────
MINIMO = 7.0   # calificación mínima aprobatoria

# ── Paleta de colores del sistema ─────────────────────────────────────────────
C_PRIMARY  = '#4f46e5'
C_SUCCESS  = '#10b981'
C_DANGER   = '#ef4444'
C_AMBER    = '#f59e0b'
C_CYAN     = '#06b6d4'
C_MUTED    = '#94a3b8'
C_TEXT     = '#1e293b'
C_TICK     = '#475569'
C_BORDER   = '#e2e8f0'
C_BG       = '#f8fafc'


# ── Helpers ───────────────────────────────────────────────────────────────────

def fig2b64(fig):
    buf = io.BytesIO()
    fig.savefig(buf, format='png', bbox_inches='tight', dpi=130, facecolor=C_BG)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()
    plt.close(fig)
    return b64


def base_ax(ax, title=None, ylabel=None, xlabel=None):
    """Apply consistent styling to any axis."""
    ax.set_facecolor(C_BG)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color(C_BORDER)
    ax.spines['bottom'].set_color(C_BORDER)
    ax.tick_params(colors=C_TICK, labelsize=8.5)
    ax.yaxis.label.set_color(C_TICK)
    ax.xaxis.label.set_color(C_TICK)
    if title:
        ax.set_title(title, color=C_TEXT, fontsize=11.5, fontweight='bold', pad=10)
    if ylabel:
        ax.set_ylabel(ylabel, color=C_TICK, fontsize=9)
    if xlabel:
        ax.set_xlabel(xlabel, color=C_TICK, fontsize=9)


def minimo_line(ax, orientation='h'):
    """Add the minimum passing line to chart."""
    if orientation == 'h':
        ax.axhline(y=MINIMO, color=C_DANGER, linestyle='--', linewidth=1.4,
                   alpha=0.8, label=f'Mínimo aprobatorio ({MINIMO:.0f})')
    else:
        ax.axvline(x=MINIMO, color=C_DANGER, linestyle='--', linewidth=1.4,
                   alpha=0.8, label=f'Mínimo ({MINIMO:.0f})')


def bar_color(val):
    return C_PRIMARY if val >= MINIMO else C_DANGER


def score_color(val):
    if val >= 8.5: return C_SUCCESS
    if val >= MINIMO: return C_CYAN
    return C_DANGER


def estilo_excel():
    return {
        'header_font': Font(bold=True, color='FFFFFF', size=11),
        'header_fill': PatternFill(start_color='3730a3', end_color='3730a3', fill_type='solid'),
        'title_fill':  PatternFill(start_color='1e1b4b', end_color='1e1b4b', fill_type='solid'),
        'aprobado_fill':   PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid'),
        'reprobado_fill':  PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid'),
        'warn_fill':       PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid'),
        'center': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin', color='d1d5db'), right=Side(style='thin', color='d1d5db'),
            top=Side(style='thin', color='d1d5db'),  bottom=Side(style='thin', color='d1d5db'),
        ),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Gráficas reutilizables (cada una tiene un propósito único)
# ─────────────────────────────────────────────────────────────────────────────

def grafica_pastel_aprobacion(aprobados, reprobados, titulo='Aprobación del grupo'):
    """Pie chart: aprobados vs reprobados."""
    fig, ax = plt.subplots(figsize=(4.8, 4), facecolor=C_BG)
    ax.set_facecolor(C_BG)
    total = aprobados + reprobados
    if total == 0:
        ax.text(0.5, 0.5, 'Sin datos', ha='center', va='center', color=C_MUTED, fontsize=12)
        ax.axis('off')
    else:
        colors = [C_SUCCESS, C_DANGER]
        explode = (0.04, 0)
        wedges, texts, autotexts = ax.pie(
            [aprobados, reprobados],
            labels=[f'Aprobados\n{aprobados}', f'Reprobados\n{reprobados}'],
            colors=colors, autopct='%1.1f%%', startangle=90,
            explode=explode, wedgeprops={'edgecolor': 'white', 'linewidth': 2},
            textprops={'fontsize': 10, 'color': C_TEXT}
        )
        for at in autotexts:
            at.set_fontweight('bold')
    ax.set_title(titulo, color=C_TEXT, fontsize=11.5, fontweight='bold', pad=8)
    plt.tight_layout()
    return fig2b64(fig)


def grafica_barras_asignaturas(nombres, promedios, titulo='Promedio por asignatura'):
    """Horizontal bar chart for subject averages."""
    n = len(nombres)
    fig, ax = plt.subplots(figsize=(8, max(3.5, n * 0.52)), facecolor=C_BG)
    base_ax(ax, titulo, xlabel='Promedio')
    colors = [bar_color(p) for p in promedios]
    bars = ax.barh(nombres, promedios, color=colors, height=0.55, edgecolor='white', linewidth=0.8)
    minimo_line(ax, 'v')
    ax.set_xlim(0, 10.5)
    ax.set_ylim(-0.5, n - 0.5)
    for bar, val in zip(bars, promedios):
        ax.text(min(val + 0.15, 10.3), bar.get_y() + bar.get_height() / 2,
                f'{val:.1f}', va='center', color=C_TEXT, fontsize=8.5, fontweight='bold')
    leg = ax.legend(loc='lower right', fontsize=8, facecolor='white',
                    edgecolor=C_BORDER, labelcolor=C_TEXT)
    plt.tight_layout()
    return fig2b64(fig)


def grafica_evolucion_parciales(promedios_por_parcial, titulo='Evolución por parcial'):
    """Line chart showing average evolution across parciales."""
    if len(promedios_por_parcial) < 2:
        return None
    keys = list(promedios_por_parcial.keys())
    vals = list(promedios_por_parcial.values())
    fig, ax = plt.subplots(figsize=(5.5, 3.8), facecolor=C_BG)
    base_ax(ax, titulo, ylabel='Promedio')
    ax.plot(keys, vals, color=C_PRIMARY, marker='o', markersize=9, linewidth=2.5,
            markerfacecolor='white', markeredgecolor=C_PRIMARY, markeredgewidth=2)
    ax.fill_between(keys, vals, alpha=0.1, color=C_PRIMARY)
    minimo_line(ax)
    ax.set_ylim(0, 10.5)
    for x, y in zip(keys, vals):
        ax.text(x, y + 0.35, f'{y:.1f}', ha='center', color=C_TEXT, fontsize=9, fontweight='bold')
    leg = ax.legend(loc='lower right', fontsize=8, facecolor='white',
                    edgecolor=C_BORDER, labelcolor=C_TEXT)
    plt.tight_layout()
    return fig2b64(fig)


def grafica_alumnos_barras(promedios_dict, titulo='Ranking de alumnos'):
    """Horizontal bar for each student's average."""
    nombres = [n[:24] for n in promedios_dict.keys()]
    promedios = list(promedios_dict.values())
    n = len(nombres)
    fig, ax = plt.subplots(figsize=(9, max(3.5, n * 0.46)), facecolor=C_BG)
    base_ax(ax, titulo, xlabel='Promedio')
    # Sort descending
    paired = sorted(zip(nombres, promedios), key=lambda x: x[1], reverse=True)
    nombres_s, promedios_s = zip(*paired) if paired else ([], [])
    colors = [bar_color(p) for p in promedios_s]
    bars = ax.barh(list(nombres_s), list(promedios_s), color=colors,
                   height=0.6, edgecolor='white', linewidth=0.8)
    minimo_line(ax, 'v')
    ax.set_xlim(0, 10.5)
    for bar, val in zip(bars, promedios_s):
        ax.text(min(val + 0.12, 10.3), bar.get_y() + bar.get_height() / 2,
                f'{val:.1f}', va='center', color=C_TEXT, fontsize=8.5, fontweight='bold')
    leg = ax.legend(loc='lower right', fontsize=8, facecolor='white',
                    edgecolor=C_BORDER, labelcolor=C_TEXT)
    plt.tight_layout()
    return fig2b64(fig)


def grafica_riesgo_asignaturas(riesgo_dict, titulo='Alumnos en riesgo por asignatura'):
    """Bar chart showing at-risk students per subject."""
    nombres = list(riesgo_dict.keys())
    vals = list(riesgo_dict.values())
    fig, ax = plt.subplots(figsize=(max(6, len(nombres) * 1.1), 4), facecolor=C_BG)
    base_ax(ax, titulo, ylabel='Alumnos en riesgo')
    colors = [C_DANGER if v > 0 else C_SUCCESS for v in vals]
    bars = ax.bar(nombres, vals, color=colors, edgecolor='white', linewidth=0.8, width=0.55)
    ax.set_ylim(0, max(vals) + 1.5 if vals else 5)
    for bar, val in zip(bars, vals):
        if val > 0:
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.05,
                    str(val), ha='center', va='bottom', color=C_TEXT, fontsize=9, fontweight='bold')
    plt.xticks(rotation=28, ha='right')
    plt.tight_layout()
    return fig2b64(fig)

def grafica_calor_asignaturas_parciales(asignaturas_resumen):
    """Heatmap-style chart: asignaturas × parciales."""
    data = []
    xlabels = ['Parcial 1', 'Parcial 2', 'Parcial 3']
    ylabels = []
    for item in asignaturas_resumen:
        ylabels.append(item['asignatura'].nombre[:18])
        row = [item['por_parcial'].get(p) for p in [1, 2, 3]]
        data.append(row)
    if not data:
        return None
    mat = np.array([[v if v is not None else np.nan for v in row] for row in data])
    fig, ax = plt.subplots(figsize=(6, max(2.5, len(ylabels) * 0.6)), facecolor=C_BG)
    base_ax(ax)
    colors = [
    (0.0, "#b91c1c"),
    (0.55, "#fa5e0a"),
    (0.6, "#f99a16fb"),   # la mitada del valor (medio transparente, puede que la barra no se vea de acuerdo al tono)
    (0.69, "#e2fb24"),   
    (0.7, "#b5f916"),    
    (1.0, "#15803d")
    ]

    cmap = LinearSegmentedColormap.from_list(
    "custom_heat",
    colors
    )

    cmap.set_bad(color='#f1f5f9')

    im = ax.imshow(
    mat,
    cmap=cmap,
    aspect='auto',
    vmin=0,
    vmax=10,
    interpolation='nearest'
    )
    cmap.set_bad(color='#f1f5f9')
    im = ax.imshow(mat, cmap=cmap, aspect='auto', vmin=0, vmax=10,
                   interpolation='nearest')
    ax.set_xticks(range(3)); ax.set_xticklabels(xlabels, fontsize=9)
    ax.set_yticks(range(len(ylabels))); ax.set_yticklabels(ylabels, fontsize=9)
    for i in range(len(ylabels)):
        for j in range(3):
            val = mat[i, j]
            if not np.isnan(val):
                color = 'white' if (val < 4 or val > 8.5) else C_TEXT
                ax.text(j, i, f'{val:.1f}', ha='center', va='center',
                        fontsize=9, fontweight='bold', color=color)
            else:
                ax.text(j, i, '—', ha='center', va='center', fontsize=9, color=C_MUTED)
    cbar = plt.colorbar(im, ax=ax, fraction=0.04, pad=0.02)
    cbar.ax.tick_params(colors=C_TICK, labelsize=8)
    ax.set_title('Mapa de calor — Promedios por asignatura y parcial',
                 color=C_TEXT, fontsize=11, fontweight='bold', pad=10)
    ax.spines[:].set_visible(False)
    plt.tight_layout()
    return fig2b64(fig)

# ─────────────────────────────────────────────────────────────────────────────
# Vistas
# ─────────────────────────────────────────────────────────────────────────────

def dashboard(request):
    semestres = Semestre.objects.all()
    grupos = Grupo.objects.all()
    total_alumnos = Alumno.objects.count()
    total_calificaciones = Calificacion.objects.count()
    califs = list(Calificacion.objects.all())
    aprobados = sum(1 for c in califs if c.aprobado)
    reprobados = len(califs) - aprobados
    pct_ap = round(aprobados / len(califs) * 100, 1) if califs else 0
    pct_rep = round(reprobados / len(califs) * 100, 1) if califs else 0
    context = {
        'semestres': semestres, 'grupos': grupos,
        'total_alumnos': total_alumnos, 'total_calificaciones': total_calificaciones,
        'aprobados': aprobados, 'reprobados': reprobados,
        'pct_aprobacion': pct_ap, 'pct_reprobacion': pct_rep,
        'minimo': MINIMO,
    }
    return render(request, 'escolar/dashboard.html', context)


def lista_alumnos(request):
    grupo_id = request.GET.get('grupo')
    grupos = Grupo.objects.select_related('semestre').all()
    alumnos = Alumno.objects.select_related('grupo__semestre').all()
    grupo_sel = None
    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        alumnos = alumnos.filter(grupo_id=grupo_id)
    return render(request, 'escolar/lista_alumnos.html',
                  {'alumnos': alumnos, 'grupos': grupos, 'grupo_sel': grupo_sel})


def calificaciones_grupo(request):
    grupo_id = request.GET.get('grupo')
    parcial = request.GET.get('parcial', '1')
    grupos = Grupo.objects.select_related('semestre').all()
    datos = []; grupo_sel = None; graficas = {}

    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        alumnos = Alumno.objects.filter(grupo=grupo_sel)
        asignaturas = Asignatura.objects.filter(semestre=grupo_sel.semestre)

        for alumno in alumnos:
            row = {'alumno': alumno, 'calificaciones': {}}
            for asig in asignaturas:
                try:
                    cal = Calificacion.objects.get(alumno=alumno, asignatura=asig, parcial=int(parcial))
                    row['calificaciones'][asig.clave] = cal
                except Calificacion.DoesNotExist:
                    row['calificaciones'][asig.clave] = None
            datos.append(row)

        # Una sola gráfica: barras de promedios por asignatura para este parcial
        promedios_asig = []
        nombres_asig = []
        for asig in asignaturas:
            califs_asig = [c for r in datos for k, c in r['calificaciones'].items() if k == asig.clave and c]
            if califs_asig:
                promedios_asig.append(round(sum(c.promedio for c in califs_asig) / len(califs_asig), 2))
                nombres_asig.append(asig.nombre[:16])

        if promedios_asig:
            graficas['asignaturas'] = grafica_barras_asignaturas(
                nombres_asig, promedios_asig,
                f'Promedio por asignatura — Parcial {parcial}'
            )

        context = {
            'grupos': grupos, 'grupo_sel': grupo_sel, 'asignaturas': asignaturas,
            'datos': datos, 'parcial': parcial, 'graficas': graficas,
            'parciales': [1, 2, 3], 'minimo': MINIMO,
        }
    else:
        context = {'grupos': grupos, 'parcial': parcial, 'parciales': [1, 2, 3], 'minimo': MINIMO}

    return render(request, 'escolar/calificaciones.html', context)


def indicadores(request):
    grupo_id = request.GET.get('grupo')
    asig_filtro = request.GET.get('asignatura', '')
    parcial_filtro = request.GET.get('parcial', '')
    grupos = Grupo.objects.select_related('semestre').all()
    graficas = {}; stats = {}; grupo_sel = None; asignaturas_resumen = []

    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        alumnos = list(Alumno.objects.filter(grupo=grupo_sel))
        asignaturas = list(Asignatura.objects.filter(semestre=grupo_sel.semestre))
        califs_all = list(Calificacion.objects.filter(alumno__grupo=grupo_sel).select_related('alumno', 'asignatura'))

        califs_f = califs_all
        if asig_filtro:
            califs_f = [c for c in califs_f if str(c.asignatura_id) == asig_filtro]
        if parcial_filtro:
            califs_f = [c for c in califs_f if str(c.parcial) == parcial_filtro]

        promedios_alumno = {}
        for alumno in alumnos:
            cal_al = [c for c in califs_f if c.alumno_id == alumno.id]
            if cal_al:
                promedios_alumno[alumno.nombre_completo] = round(sum(c.promedio for c in cal_al) / len(cal_al), 2)

        stats['total'] = len(alumnos)
        stats['aprobados'] = sum(1 for p in promedios_alumno.values() if p >= MINIMO)
        stats['reprobados'] = sum(1 for p in promedios_alumno.values() if p < MINIMO)
        stats['pct_aprobacion'] = round(stats['aprobados'] / stats['total'] * 100, 1) if stats['total'] else 0
        stats['pct_reprobacion'] = round(stats['reprobados'] / stats['total'] * 100, 1) if stats['total'] else 0
        stats['promedio_grupo'] = round(sum(promedios_alumno.values()) / len(promedios_alumno), 2) if promedios_alumno else 0

        for asig in asignaturas:
            if asig_filtro and str(asig.id) != asig_filtro:
                continue
            cals = [c for c in califs_all if c.asignatura_id == asig.id]
            if parcial_filtro:
                cals = [c for c in cals if str(c.parcial) == parcial_filtro]
            if not cals:
                continue
            ap = sum(1 for c in cals if c.aprobado)
            rep = len(cals) - ap
            prom = round(sum(c.promedio for c in cals) / len(cals), 2)
            por_parcial = {}
            for p in [1, 2, 3]:
                cp = [c for c in cals if c.parcial == p]
                if cp:
                    por_parcial[p] = round(sum(c.promedio for c in cp) / len(cp), 2)
            asignaturas_resumen.append({
                'asignatura': asig, 'total': len(set(c.alumno_id for c in cals)),
                'aprobados': ap, 'reprobados': rep, 'promedio': prom,
                'pct_aprobacion': round(ap / len(cals) * 100, 1) if cals else 0,
                'por_parcial': por_parcial,
            })

        # Gráfica 1: Pastel aprobación (solo si hay datos)
        if stats['total']:
            graficas['pastel'] = grafica_pastel_aprobacion(stats['aprobados'], stats['reprobados'])

        # Gráfica 2: Evolución parciales (solo si no hay filtro de parcial)
        if not parcial_filtro:
            promedios_parcial = {}
            for p in [1, 2, 3]:
                cp = [c for c in califs_f if c.parcial == p]
                if cp:
                    promedios_parcial[f'P{p}'] = round(sum(c.promedio for c in cp) / len(cp), 2)
            g_evo = grafica_evolucion_parciales(promedios_parcial)
            if g_evo:
                graficas['evolucion'] = g_evo

        # Gráfica 3: Ranking de alumnos (ordenado)
        if promedios_alumno:
            graficas['alumnos'] = grafica_alumnos_barras(promedios_alumno)

        # Gráfica 4: Mapa de calor asignaturas × parciales (solo sin filtros activos)
        if asignaturas_resumen and not parcial_filtro and not asig_filtro:
            g_calor = grafica_calor_asignaturas_parciales(asignaturas_resumen)
            if g_calor:
                graficas['calor'] = g_calor

    context = {
        'grupos': grupos, 'grupo_sel': grupo_sel, 'stats': stats, 'graficas': graficas,
        'asignaturas_resumen': asignaturas_resumen,
        'asignaturas_lista': list(Asignatura.objects.filter(semestre=grupo_sel.semestre)) if grupo_sel else [],
        'asig_filtro': asig_filtro, 'parcial_filtro': parcial_filtro, 'minimo': MINIMO,
    }
    return render(request, 'escolar/indicadores.html', context)


def pronostico_ia(request):
    grupo_id = request.GET.get('grupo')
    grupos = Grupo.objects.select_related('semestre').all()
    resultado = []; grupo_sel = None; graficas = {}
    indices_asig = []; stats_grupo = {}

    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        alumnos = list(Alumno.objects.filter(grupo=grupo_sel).order_by('nombre_completo'))
        asignaturas = list(Asignatura.objects.filter(semestre=grupo_sel.semestre))
        califs_all = list(Calificacion.objects.filter(alumno__grupo=grupo_sel).select_related('alumno', 'asignatura'))

        for alumno in alumnos:
            datos_alumno = {'alumno': alumno, 'asignaturas': [], 'promedio_global': None}
            promedios_g = []
            for asig in asignaturas:
                califs = [c for c in califs_all if c.alumno_id == alumno.id and c.asignatura_id == asig.id]
                promedios = [c.promedio for c in sorted(califs, key=lambda x: x.parcial)]
                if len(promedios) >= 2:
                    x = np.array(range(1, len(promedios) + 1))
                    coef = np.polyfit(x, np.array(promedios), 1)
                    pron = round(max(0, min(10, float(np.polyval(coef, 3)))), 2)
                    tendencia = 'sube' if coef[0] > 0.2 else ('baja' if coef[0] < -0.2 else 'estable')
                elif len(promedios) == 1:
                    pron = promedios[0]; tendencia = 'sin datos'
                else:
                    pron = None; tendencia = 'sin datos'
                if pron is not None:
                    promedios_g.append(pron)
                datos_alumno['asignaturas'].append({
                    'asignatura': asig, 'promedios': promedios, 'pronostico': pron,
                    'tendencia': tendencia, 'riesgo': pron is not None and pron < MINIMO,
                })
            if promedios_g:
                datos_alumno['promedio_global'] = round(sum(promedios_g) / len(promedios_g), 2)
            resultado.append(datos_alumno)

        pgs = [d['promedio_global'] for d in resultado if d['promedio_global'] is not None]
        stats_grupo = {
            'aprobados': sum(1 for p in pgs if p >= MINIMO),
            'reprobados': sum(1 for p in pgs if p < MINIMO),
            'total': len(alumnos),
        }
        if stats_grupo['total']:
            stats_grupo['pct_aprobacion'] = round(stats_grupo['aprobados'] / stats_grupo['total'] * 100, 1)
            stats_grupo['pct_reprobacion'] = round(stats_grupo['reprobados'] / stats_grupo['total'] * 100, 1)

        for asig in asignaturas:
            prons = [
                a['pronostico'] for d in resultado
                for a in d['asignaturas']
                if a['asignatura'].id == asig.id and a['pronostico'] is not None
            ]
            ap = sum(1 for p in prons if p >= MINIMO)
            rep = len(prons) - ap
            indices_asig.append({
                'asignatura': asig, 'aprobados': ap, 'reprobados': rep, 'total': len(prons),
                'pct_aprobacion': round(ap / len(prons) * 100, 1) if prons else 0,
                'promedio': round(sum(prons) / len(prons), 2) if prons else 0,
            })

        # Gráfica 1: Pastel grupo (pronóstico)
        if stats_grupo.get('total'):
            graficas['pastel'] = grafica_pastel_aprobacion(
                stats_grupo['aprobados'], stats_grupo['reprobados'], 'Pronóstico del grupo')

        # Gráfica 2: Alumnos en riesgo por asignatura
        riesgo = {
            asig.nombre[:16]: sum(1 for d in resultado for a in d['asignaturas']
                                   if a['asignatura'].id == asig.id and a['riesgo'])
            for asig in asignaturas
        }
        if any(v > 0 for v in riesgo.values()):
            graficas['riesgo'] = grafica_riesgo_asignaturas(riesgo)

    context = {
        'grupos': grupos, 'grupo_sel': grupo_sel, 'resultado': resultado, 'graficas': graficas,
        'indices_asig': indices_asig, 'stats_grupo': stats_grupo, 'minimo': MINIMO,
    }
    return render(request, 'escolar/pronostico.html', context)


# ── Exportar Excel ─────────────────────────────────────────────────────────────

def exportar_excel(request):
    grupo_id = request.GET.get('grupo')
    parcial = request.GET.get('parcial', 'todos')
    if not grupo_id:
        messages.error(request, 'Selecciona un grupo')
        return redirect('lista_alumnos')
    grupo = get_object_or_404(Grupo, id=grupo_id)
    alumnos = Alumno.objects.filter(grupo=grupo).order_by('nombre_completo')
    asignaturas = Asignatura.objects.filter(semestre=grupo.semestre)
    e = estilo_excel()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    parciales_a = [1, 2, 3] if parcial == 'todos' else [int(parcial)]

    for p in parciales_a:
        ws = wb.create_sheet(title=f'Parcial {p}')
        ws.merge_cells('A1:H1')
        ws['A1'] = f'TRAYECTORIA ESCOLAR — {grupo.semestre.nombre} — Grupo {grupo.nombre} — Parcial {p}  (Mín. aprobatorio: {MINIMO:.0f})'
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
        ws['A1'].fill = e['title_fill']; ws['A1'].alignment = e['center']
        ws.row_dimensions[1].height = 26
        for col, h in enumerate(['N°', 'No. Cuenta', 'Nombre', 'Asignatura', 'Hetero', 'Coevaluación', 'Autoevaluación', 'Promedio'], 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = e['header_font']; c.fill = e['header_fill']
            c.alignment = e['center']; c.border = e['border']
        for i, w in enumerate([5, 15, 32, 28, 12, 14, 14, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        row_n = 3
        for idx, alumno in enumerate(alumnos, 1):
            for asig in asignaturas:
                try:
                    cal = Calificacion.objects.get(alumno=alumno, asignatura=asig, parcial=p)
                    fill_r = e['aprobado_fill'] if cal.aprobado else e['reprobado_fill']
                    row_d = [idx, alumno.numero_cuenta, alumno.nombre_completo, asig.nombre,
                             float(cal.heteroevaluacion), float(cal.coevaluacion), float(cal.autoevaluacion), cal.promedio]
                except Calificacion.DoesNotExist:
                    fill_r = e['warn_fill']
                    row_d = [idx, alumno.numero_cuenta, alumno.nombre_completo, asig.nombre, 'N/A', 'N/A', 'N/A', 'N/A']
                for col, val in enumerate(row_d, 1):
                    c = ws.cell(row=row_n, column=col, value=val)
                    c.fill = fill_r; c.border = e['border']
                    c.alignment = e['center'] if col in [1, 2, 5, 6, 7, 8] else Alignment(vertical='center')
                row_n += 1

    ws_lista = wb.create_sheet(title='Lista Original')
    ws_lista.merge_cells('A1:D1')
    ws_lista['A1'] = f'LISTA DE ALUMNOS — {grupo.semestre.nombre} — Grupo {grupo.nombre}'
    ws_lista['A1'].font = Font(bold=True, color='FFFFFF', size=12)
    ws_lista['A1'].fill = e['title_fill']; ws_lista['A1'].alignment = e['center']
    ws_lista.row_dimensions[1].height = 24
    for col, h in enumerate(['N°', 'No. Cuenta', 'Nombre Completo', 'Correo'], 1):
        c = ws_lista.cell(row=2, column=col, value=h)
        c.font = e['header_font']; c.fill = e['header_fill']
        c.alignment = e['center']; c.border = e['border']
    for col_w, w in zip(['A', 'B', 'C', 'D'], [5, 15, 40, 35]):
        ws_lista.column_dimensions[col_w].width = w
    alt = PatternFill(start_color='eef2ff', end_color='eef2ff', fill_type='solid')
    for i, alumno in enumerate(alumnos, 1):
        fill_r = alt if i % 2 == 0 else PatternFill(fill_type=None)
        for col, val in enumerate([i, alumno.numero_cuenta, alumno.nombre_completo, alumno.email], 1):
            c = ws_lista.cell(row=i + 2, column=col, value=val)
            c.fill = fill_r; c.border = e['border']
            c.alignment = e['center'] if col in [1, 2] else Alignment(vertical='center')

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="trayectoria_{grupo.semestre.nombre.replace(" ","_")}_G{grupo.nombre}_P{parcial}.xlsx"'
    wb.save(resp)
    return resp


def exportar_pronostico_excel(request):
    grupo_id = request.GET.get('grupo')
    if not grupo_id:
        return redirect('pronostico')
    grupo = get_object_or_404(Grupo, id=grupo_id)
    alumnos = list(Alumno.objects.filter(grupo=grupo).order_by('nombre_completo'))
    asignaturas = list(Asignatura.objects.filter(semestre=grupo.semestre))
    califs_all = list(Calificacion.objects.filter(alumno__grupo=grupo).select_related('alumno', 'asignatura'))
    e = estilo_excel()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet('Pronóstico General')
    ws.merge_cells('A1:G1')
    ws['A1'] = f'PRONÓSTICO IA — {grupo.semestre.nombre} — Grupo {grupo.nombre}  (Mín. aprobatorio: {MINIMO:.0f})'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
    ws['A1'].fill = e['title_fill']; ws['A1'].alignment = e['center']
    ws.row_dimensions[1].height = 26
    for col, h in enumerate(['Alumno', 'No. Cuenta', 'P1', 'P2', 'P3', 'Pronóstico P3 (IA)', 'Tendencia'], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = e['header_font']; c.fill = e['header_fill']
        c.alignment = e['center']; c.border = e['border']
    for i, w in enumerate([35, 15, 10, 10, 10, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    row_n = 3; ap_g = 0; rep_g = 0
    for alumno in alumnos:
        cal_al = [c for c in califs_all if c.alumno_id == alumno.id]
        pp = {p: round(sum(c.promedio for c in cal_al if c.parcial == p) / len([c for c in cal_al if c.parcial == p]), 2)
              for p in [1, 2, 3] if [c for c in cal_al if c.parcial == p]}
        vals = list(pp.values())
        if len(vals) >= 2:
            coef = np.polyfit(range(1, len(vals)+1), vals, 1)
            pron = round(max(0, min(10, float(np.polyval(coef, 3)))), 2)
            tend = 'Sube' if coef[0] > 0.2 else ('Baja' if coef[0] < -0.2 else 'Estable')
        elif len(vals) == 1:
            pron = vals[0]; tend = 'Sin datos'
        else:
            pron = None; tend = '—'
        if pron is not None and pron >= MINIMO: ap_g += 1
        elif pron is not None: rep_g += 1
        fill_r = e['aprobado_fill'] if (pron and pron >= MINIMO) else e['reprobado_fill']
        for col, val in enumerate([alumno.nombre_completo, alumno.numero_cuenta,
                                    pp.get(1,'N/A'), pp.get(2,'N/A'), pp.get(3,'N/A'),
                                    pron if pron else '—', tend], 1):
            c = ws.cell(row=row_n, column=col, value=val)
            c.fill = fill_r; c.border = e['border']; c.alignment = e['center']
        row_n += 1
    row_n += 1
    ws.cell(row=row_n, column=1, value=f'Aprobados (pronóstico): {ap_g}  |  Reprobados: {rep_g}  |  % Aprobación: {round(ap_g/len(alumnos)*100,1) if alumnos else 0}%').font = Font(bold=True, size=10)

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="pronostico_IA_{grupo.semestre.nombre.replace(" ","_")}_G{grupo.nombre}.xlsx"'
    wb.save(resp)
    return resp


# ── Exportar por asignatura ───────────────────────────────────────────────────

def exportar_asignatura(request):
    asig_id = request.GET.get('asignatura')
    parcial = request.GET.get('parcial', 'todos')
    grupo_id = request.GET.get('grupo')
    if not asig_id:
        return redirect('profesor')
    asig = get_object_or_404(Asignatura, id=asig_id)
    e = estilo_excel()
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    alumnos_qs = Alumno.objects.filter(grupo__semestre=asig.semestre)
    if grupo_id:
        alumnos_qs = alumnos_qs.filter(grupo_id=grupo_id)
    alumnos = list(alumnos_qs.order_by('nombre_completo'))
    for p in ([1, 2, 3] if parcial == 'todos' else [int(parcial)]):
        ws = wb.create_sheet(title=f'Parcial {p}')
        ws.merge_cells('A1:H1')
        ws['A1'] = f'{asig.nombre} ({asig.clave}) — Parcial {p}  (Mín.: {MINIMO:.0f})'
        ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
        ws['A1'].fill = e['title_fill']; ws['A1'].alignment = e['center']
        ws.row_dimensions[1].height = 26
        for col, h in enumerate(['N°', 'No. Cuenta', 'Nombre', 'Grupo', 'Hetero', 'Coevaluación', 'Autoevaluación', 'Promedio'], 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = e['header_font']; c.fill = e['header_fill']
            c.alignment = e['center']; c.border = e['border']
        for i, w in enumerate([5, 15, 35, 10, 12, 14, 14, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for idx, alumno in enumerate(alumnos, 1):
            try:
                cal = Calificacion.objects.get(alumno=alumno, asignatura=asig, parcial=p)
                fill_r = e['aprobado_fill'] if cal.aprobado else e['reprobado_fill']
                row_d = [idx, alumno.numero_cuenta, alumno.nombre_completo, alumno.grupo.nombre,
                         float(cal.heteroevaluacion), float(cal.coevaluacion), float(cal.autoevaluacion), cal.promedio]
            except Calificacion.DoesNotExist:
                fill_r = e['warn_fill']
                row_d = [idx, alumno.numero_cuenta, alumno.nombre_completo, alumno.grupo.nombre, 'N/A', 'N/A', 'N/A', 'N/A']
            for col, val in enumerate(row_d, 1):
                c = ws.cell(row=idx + 2, column=col, value=val)
                c.fill = fill_r; c.border = e['border']; c.alignment = e['center']

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="asignatura_{asig.clave}_P{parcial}.xlsx"'
    wb.save(resp)
    return resp


def descargar_plantilla(request):
    e = estilo_excel()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'Calificaciones'
    ws.merge_cells('A1:E1')
    ws['A1'] = f'PLANTILLA DE IMPORTACIÓN — Trayectoria Escolar (Mín. aprobatorio: {MINIMO:.0f})'
    ws['A1'].font = Font(bold=True, color='FFFFFF', size=12)
    ws['A1'].fill = e['title_fill']; ws['A1'].alignment = e['center']
    ws.row_dimensions[1].height = 26
    ws.merge_cells('A2:E2')
    ws['A2'] = 'Instrucciones: Llenar desde la fila 4. Calificaciones en escala 0-10. NO modificar encabezados.'
    ws['A2'].font = Font(italic=True, color='888888', size=10)
    ws['A2'].alignment = Alignment(horizontal='center')
    for col, h in enumerate(['Número de Cuenta', 'Nombre Completo', 'Heteroevaluación', 'Coevaluación', 'Autoevaluación'], 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = e['header_font']; c.fill = e['header_fill']
        c.alignment = e['center']; c.border = e['border']
    ejemplos = [('2024001','García López María',8.5,7.0,9.0),('2024002','Hernández Juan',6.0,7.5,7.5),('2024003','Martínez Torres Ana',9.0,8.0,8.5)]
    alt = PatternFill(start_color='eef2ff', end_color='eef2ff', fill_type='solid')
    for i, (nc, nm, h, c, a) in enumerate(ejemplos, 4):
        fill_r = alt if i % 2 == 0 else PatternFill(fill_type=None)
        for col, val in enumerate([nc, nm, h, c, a], 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = fill_r; cell.border = e['border']; cell.alignment = e['center']
    for i, w in enumerate([20, 40, 18, 15, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="plantilla_calificaciones.xlsx"'
    wb.save(resp); return resp


# ── Captura rápida ────────────────────────────────────────────────────────────

def captura_rapida(request):
    grupos = Grupo.objects.select_related('semestre').all()
    grupo_id = request.GET.get('grupo')
    asig_id = request.GET.get('asignatura')
    parcial = request.GET.get('parcial', '1')
    grupo_sel = asig_sel = None
    alumnos = []; asignaturas = []
    alumno_ids_json = '[]'

    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        asignaturas = list(Asignatura.objects.filter(semestre=grupo_sel.semestre))
        if asig_id:
            asig_sel = get_object_or_404(Asignatura, id=asig_id)
            for al in Alumno.objects.filter(grupo=grupo_sel).order_by('nombre_completo'):
                try:
                    al.cal_existente = Calificacion.objects.get(alumno=al, asignatura=asig_sel, parcial=int(parcial))
                except Calificacion.DoesNotExist:
                    al.cal_existente = None
                alumnos.append(al)
            alumno_ids_json = _json.dumps([a.id for a in alumnos])

    return render(request, 'escolar/captura_rapida.html', {
        'grupos': grupos, 'grupo_sel': grupo_sel, 'asig_sel': asig_sel,
        'asignaturas': asignaturas, 'alumnos': alumnos, 'parcial': parcial,
        'alumno_ids_json': alumno_ids_json, 'minimo': MINIMO,
    })


def guardar_calificaciones(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Método no permitido'})
    try:
        data = _json.loads(request.body)
        asig = get_object_or_404(Asignatura, id=data['asignatura_id'])
        parcial = int(data['parcial'])
        creados = actualizados = 0
        for row in data.get('rows', []):
            alumno = get_object_or_404(Alumno, id=row['alumno_id'])
            h, c, a = float(row['hetero']), float(row['co']), float(row['auto'])
            if not (0 <= h <= 10 and 0 <= c <= 10 and 0 <= a <= 10):
                continue
            _, created = Calificacion.objects.update_or_create(
                alumno=alumno, asignatura=asig, parcial=parcial,
                defaults={'heteroevaluacion': h, 'coevaluacion': c, 'autoevaluacion': a}
            )
            if created: creados += 1
            else: actualizados += 1
        return JsonResponse({'ok': True, 'creados': creados, 'actualizados': actualizados})
    except Exception as ex:
        return JsonResponse({'ok': False, 'error': str(ex)})


# ── Seguimiento ───────────────────────────────────────────────────────────────

def seguimiento(request):
    from .models import Nota
    grupos = Grupo.objects.select_related('semestre').all()
    grupo_id = request.GET.get('grupo')
    grupo_sel = None
    alumnos = list(Alumno.objects.select_related('grupo').prefetch_related('notas', 'calificaciones').all())
    if grupo_id:
        grupo_sel = get_object_or_404(Grupo, id=grupo_id)
        alumnos = [a for a in alumnos if a.grupo_id == int(grupo_id)]
    for al in alumnos:
        cals = list(al.calificaciones.all())
        al.promedio_general = round(sum(c.promedio for c in cals) / len(cals), 2) if cals else None
    return render(request, 'escolar/seguimiento.html',
                  {'grupos': grupos, 'grupo_sel': grupo_sel, 'alumnos': alumnos, 'minimo': MINIMO})


def seguimiento_notas_api(request):
    from .models import Nota
    alumno = get_object_or_404(Alumno, id=request.GET.get('alumno_id'))
    cals = list(alumno.calificaciones.all())
    prom = round(sum(c.promedio for c in cals) / len(cals), 2) if cals else None
    notas = list(Nota.objects.filter(alumno=alumno).values('tipo', 'texto', 'fecha'))
    for n in notas:
        n['fecha'] = n['fecha'].strftime('%d/%m/%Y %H:%M')
    return JsonResponse({'ok': True, 'notas': notas, 'promedio': prom})


def agregar_nota_api(request):
    from .models import Nota
    if request.method != 'POST':
        return JsonResponse({'ok': False})
    data = _json.loads(request.body)
    alumno = get_object_or_404(Alumno, id=data['alumno_id'])
    Nota.objects.create(alumno=alumno, tipo=data.get('tipo', 'observacion'), texto=data.get('texto', ''))
    return JsonResponse({'ok': True})


# ── Panel Profesor ────────────────────────────────────────────────────────────

def profesor_panel(request):
    pasos_flujo = [
        ('Descarga la plantilla Excel', 'Formato predefinido con las columnas correctas'),
        ('Llena con los datos del grupo', 'Número de cuenta, nombre y las 3 evaluaciones'),
        ('Selecciona grupo, asignatura y parcial', 'El sistema sabe a qué registro corresponde'),
        ('Haz clic en Importar', 'Los alumnos y calificaciones se crean automáticamente'),
    ]
    return render(request, 'escolar/profesor.html', {
        'semestres': Semestre.objects.all(),
        'grupos': Grupo.objects.select_related('semestre').all(),
        'asignaturas': Asignatura.objects.select_related('semestre').all(),
        'pasos_flujo': pasos_flujo,
        'minimo': MINIMO,
    })


def crear_estructura(request):
    if request.method != 'POST':
        return redirect('profesor')
    accion = request.POST.get('accion')
    if accion == 'semestre':
        nombre = request.POST.get('nombre_semestre', '').strip()
        anio = request.POST.get('anio', '')
        periodo = request.POST.get('periodo', '')
        if nombre and anio and periodo:
            sem, created = Semestre.objects.get_or_create(nombre=nombre, anio=int(anio), periodo=periodo)
            messages.success(request, f'Semestre "{sem.nombre}" {"creado" if created else "ya existía"}.')
        else:
            messages.error(request, 'Completa todos los campos del semestre.')
    elif accion == 'grupo':
        sem_id = request.POST.get('semestre_grupo')
        nombre_g = request.POST.get('nombre_grupo', '').strip()
        if sem_id and nombre_g:
            sem = get_object_or_404(Semestre, id=sem_id)
            g, created = Grupo.objects.get_or_create(nombre=nombre_g, semestre=sem)
            messages.success(request, f'Grupo "{g.nombre}" {"creado" if created else "ya existía"}.')
        else:
            messages.error(request, 'Selecciona semestre y nombre de grupo.')
    elif accion == 'asignatura':
        sem_id = request.POST.get('semestre_asig')
        nombre_a = request.POST.get('nombre_asig', '').strip()
        clave_a = request.POST.get('clave_asig', '').strip().upper()
        if sem_id and nombre_a and clave_a:
            sem = get_object_or_404(Semestre, id=sem_id)
            a, created = Asignatura.objects.get_or_create(clave=clave_a, defaults={'nombre': nombre_a, 'semestre': sem})
            messages.success(request, f'Asignatura "{a.nombre}" ({a.clave}) {"creada" if created else "ya existía"}.')
        else:
            messages.error(request, 'Completa todos los campos de la asignatura.')
    elif accion == 'alumno':
        num_cuenta = request.POST.get('numero_cuenta', '').strip()
        nombre = request.POST.get('nombre_alumno', '').strip()
        email = request.POST.get('email_alumno', '').strip()
        grupo_id = request.POST.get('grupo_alumno')
        if num_cuenta and nombre and grupo_id:
            grupo = get_object_or_404(Grupo, id=grupo_id)
            al, created = Alumno.objects.get_or_create(
                numero_cuenta=num_cuenta,
                defaults={'nombre_completo': nombre, 'email': email or f'{num_cuenta}@ithi.edu.mx', 'grupo': grupo}
            )
            messages.success(request, f'Alumno "{al.nombre_completo}" {"agregado" if created else "ya estaba registrado"}.')
        else:
            messages.error(request, 'Completa número de cuenta, nombre y grupo.')
    return redirect('profesor')


def importar_excel(request):
    if request.method != 'POST':
        return redirect('profesor')
    archivo = request.FILES.get('archivo_excel')
    grupo_id = request.POST.get('grupo_importar')
    asig_id = request.POST.get('asignatura_importar')
    parcial = request.POST.get('parcial_importar')
    if not all([archivo, grupo_id, asig_id, parcial]):
        messages.error(request, 'Completa todos los campos antes de importar.')
        return redirect('profesor')
    grupo = get_object_or_404(Grupo, id=grupo_id)
    asig = get_object_or_404(Asignatura, id=asig_id)
    try:
        wb = openpyxl.load_workbook(archivo); ws = wb.active
        creados = actualizados = 0; errores = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not row or not row[0]:
                continue
            try:
                nc = str(row[0]).strip()
                nombre = str(row[1]).strip() if row[1] else ''
                h, c, a = float(row[2]), float(row[3]), float(row[4])
                if not (0 <= h <= 10 and 0 <= c <= 10 and 0 <= a <= 10):
                    errores.append(f'Fila {row_idx}: calificaciones fuera de rango'); continue
                alumno, al_c = Alumno.objects.get_or_create(
                    numero_cuenta=nc,
                    defaults={'nombre_completo': nombre, 'email': f'{nc}@ithi.edu.mx', 'grupo': grupo}
                )
                _, cal_c = Calificacion.objects.update_or_create(
                    alumno=alumno, asignatura=asig, parcial=int(parcial),
                    defaults={'heteroevaluacion': h, 'coevaluacion': c, 'autoevaluacion': a}
                )
                if cal_c: creados += 1
                else: actualizados += 1
            except Exception as ex:
                errores.append(f'Fila {row_idx}: {ex}')
        msg = f'Importación completa: {creados} nuevos, {actualizados} actualizados.'
        if errores:
            msg += f' {len(errores)} errores.'
            messages.warning(request, msg)
        else:
            messages.success(request, msg)
    except Exception as ex:
        messages.error(request, f'Error al leer el archivo: {ex}')
    return redirect('profesor')
